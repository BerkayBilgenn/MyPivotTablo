import os
import io
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS  # CORS desteği eklendi
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt

app = Flask(__name__)
CORS(app)  # Tüm isteklerde CORS hatasını önler

# Ana sayfa (HTML yükler)
@app.route('/')
def index():
    return render_template('index.html')

# Excel dosyasını yüklemek ve verileri işlemek
@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({"success": False, "message": "Dosya bulunamadı"}), 400

        # Excel dosyasını oku
        df = pd.read_excel(file)

        # Sütun adlarını normalize et (küçük harf yap ve boşlukları kaldır)
        df.columns = df.columns.str.strip().str.lower()

        # Hedef sütunları bulmak için anahtar kelimeler
        x_column = next((col for col in df.columns if "çağrı" in col or "rapor" in col), None)
        toplam_gelen_column = next((col for col in df.columns if "gelen" in col), None)
        cevaplanan_column = next((col for col in df.columns if "cevaplanan" in col), None)

        # Gerekli sütunlar yoksa hata döndür
        if not x_column or not toplam_gelen_column or not cevaplanan_column:
            return jsonify({"success": False, "message": "Gerekli sütunlar bulunamadı!"}), 400

        # Verileri işle
        x_coords = df[x_column].fillna("Bilinmiyor").astype(str).tolist()
        toplam_gelen = df[toplam_gelen_column].fillna(0).astype(int).tolist()
        toplam_cevaplanan = df[cevaplanan_column].fillna(0).astype(int).tolist()

        return jsonify({
            "success": True,
            "data": {
                "x_coords": x_coords,
                "toplam_gelen": toplam_gelen,
                "toplam_cevaplanan": toplam_cevaplanan
            }
        })
    except Exception as e:
        return jsonify({"success": False, "message": f"Hata: {str(e)}"}), 500

# Excel ve grafik oluşturma
@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "Geçerli veri gönderilmedi"}), 400

        x_coords = data.get('x_coords', [])
        cevaplanan = data.get('cevaplanan', [])
        gelen = data.get('gelen', [])

        if not x_coords or not cevaplanan or not gelen:
            return jsonify({"success": False, "message": "Eksik veri gönderildi"}), 400

        df = pd.DataFrame({
            'X Koordinatları': x_coords,
            'Toplam Cevaplanan Çağrı': cevaplanan,
            'Toplam Gelen Çağrı': gelen
        })

        # Cevaplanan / Gelen Oranı Hesapla
        df['Cevaplanan / Gelen Çağrı Oranı (%)'] = (
            df['Toplam Cevaplanan Çağrı'] / df['Toplam Gelen Çağrı'].replace(0, 1)
        ) * 100

        # Grafik Oluştur
        plt.figure(figsize=(14, 8))
        plt.bar(df['X Koordinatları'], df['Toplam Gelen Çağrı'], label='Toplam Gelen Çağrı', color='cyan')
        plt.bar(df['X Koordinatları'], df['Toplam Cevaplanan Çağrı'], label='Toplam Cevaplanan Çağrı', color='violet')
        plt.plot(df['X Koordinatları'], df['Cevaplanan / Gelen Çağrı Oranı (%)'], label='Cevaplanan / Gelen Oranı (%)', color='orange', marker='o', linewidth=2)
        plt.xlabel('X Koordinatları')
        plt.ylabel('Çağrı Sayısı')
        plt.title('Çağrı Verileri')
        plt.legend()
        plt.tight_layout()

        # Grafiği kaydet
        chart_image = io.BytesIO()
        plt.savefig(chart_image, format='png')
        plt.close()
        chart_image.seek(0)

        # Excel Dosyası Oluştur
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Veriler"

        # Verileri Excel'e Aktar
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Tablo Stilini Uygula
        tab = Table(displayName="CallData", ref=f"A1:E{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Excel Grafiği Ekle
        chart = BarChart()
        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = "Çağrı Verileri"
        chart.style = 10
        ws.add_chart(chart, "G2")

        # Grafik Sayfası
        chart_ws = wb.create_sheet(title="Grafik")
        img = Image(chart_image)
        img.anchor = 'A1'
        chart_ws.add_image(img)

        # Dosyayı kaydet ve gönder
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='veriler_pivot_grafikli.xlsx'
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Hata: {str(e)}"}), 500

# Flask Sunucusunu Başlat
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
